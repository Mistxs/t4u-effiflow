import datetime
import threading

from flask import jsonify, request, Blueprint, send_file, make_response
import requests
import json
from russiannames.parser import NamesParser
from openpyxl import Workbook
import io
from config import bearer, db_params
from concurrent.futures import ThreadPoolExecutor
from apps.more.socketio_setup import socketio


import pymysql


fiosplitter = Blueprint('fiosplitter', __name__)


connection = pymysql.connect(**db_params)


headers = {
    'Content-Type': 'application/json',
    'Authorization': f'{bearer}',
    'Accept': 'application/vnd.yclients.v2+json'
}


def create_table():
    conn = pymysql.connect(**db_params)
    cursor = conn.cursor()
    cursor.execute(
        'CREATE TABLE IF NOT EXISTS fiosplitter (id INT AUTO_INCREMENT PRIMARY KEY, salon_id INT, data TEXT);'
    )
    cursor.execute(
        'CREATE TABLE IF NOT EXISTS clients(id int, salon_id int, fullname varchar(255), phone varchar(100), name varchar(255), surname varchar(255), patronymic varchar(255), birthday varchar(255), query_ts datetime);'
    )

    conn.commit()
    conn.close()


create_table()


def safeClients(salon, data):
    try:
        conn = pymysql.connect(**db_params)
        cursor = conn.cursor()
        current_time = datetime.datetime.now()

        for item in data:
            query = """
                       INSERT INTO clients (yc_id, salon_id, phone, name, surname, patronymic, birthday, query_ts)
                       VALUES (%s, %s, %s, %s, %s, %s, %s, %s)
                   """
            clid = item.get("id")
            phone = item.get("phone")
            name = item.get("name")
            surname = item.get("surname")
            patronymic = item.get("patronymic")
            birthday = item.get("birthday")



            values = (clid, int(salon),  phone, name, surname, patronymic, birthday, current_time)
            cursor.execute(query, values)

        conn.commit()
        conn.close()

        print("Data successfully saved in the 'clients' table.")
    except Exception as e:
        print(f"Error: {e}")


clientsdata = []
newdata = []
errors = []


@fiosplitter.route('/getClients', methods=['POST'])
def getcli():
    try:
        response = request.json
        salon = response["salon_id"]
        login = response["login"]
        password = response["password"]

        if password == "":
            token = login
            clientsdata = getClients(salon, headers, token)
            getBirthDay(salon, headers, token)
        else:
            usertoken = getToken(login, password)
            clientsdata = getClients(salon, headers, usertoken)
            getBirthDay(salon, headers, usertoken)
        return jsonify({"status": "success", "text": clientsdata})
    except Exception as e:
        return jsonify({'status': 'error', 'text': f'{e}'})


def getToken(login, password):
    url = f"https://api.yclients.com/api/v1/auth"
    payload = json.dumps({
        "login": f"{login}",
        "password": f"{password}"
    })
    response = requests.request("POST", url, headers=headers, data=payload).json()
    if response["success"] == True:
        return response["data"]["user_token"]
    else:
        return f"Ошибка. {response['meta']}"


def getClients(salon, headers, token):
    page = 1
    global clientsdata
    clientsdata = []
    headers['Authorization'] = f'{bearer}, {token}'
    totalcount = -1

    while totalcount == -1 or page <= totalcount / 200 + 1:
        url = f"https://api.yclients.com/api/v1/company/{salon}/clients/search"
        payload = json.dumps({
            "page": page,
            "page_size": 200,
            "fields": [
                "id",
                "phone",
                "name",
                "surname",
                "patronymic"
            ]
        })
        response = requests.request("POST", url, headers=headers, data=payload).json()
        if totalcount == -1:
            totalcount = response["meta"]["total_count"]

        for item in response["data"]:
            clientsdata.append({
                "id": item.get("id", ""),
                "phone": item.get("phone", ""),
                "name": item.get("name", ""),
                "patronymic": item.get("patronymic", ""),
                "surname": item.get("surname", "")
            })

        socketio.emit('parseclients', {'counter': {'totalclients':totalcount, 'currentcount':len(clientsdata)}})
        page += 1
    return clientsdata


def getBirthDay(salon, headers, token):
    page = 1
    headers['Authorization'] = f'{bearer}, {token}'
    totalcount = -1
    currentcount = 0
    while totalcount == -1 or page <= totalcount / 300 + 1:
        url = f"https://api.yclients.com/api/v1/clients/{salon}?count=300&page={page}"
        response = requests.request("GET", url, headers=headers).json()
        if totalcount == -1:
            totalcount = response["meta"]["total_count"]
        for item in response["data"]:
            for client in clientsdata:
                if client["id"] == item["id"]:
                    client["birthday"] = item.get("birth_date", "")
                    currentcount += 1
                    socketio.emit('addedBD',
                                  {'counter': {'totalclients': totalcount, 'currentcount': currentcount}})
                    break
        page += 1

    safeClients(salon, clientsdata)
    return clientsdata


@fiosplitter.route('/parsefio', methods=['POST'])
def parsefio():
    try:
        global newdata
        response = request.json
        salon = response["salon_id"]
        login = response["login"]
        password = response["password"]
        newdata = parseclients(clientsdata)
        print(newdata)
        # newdata = splitByChar(clientsdata)
        return jsonify({'status': 'success', 'text': newdata})
    except Exception as e:
        return jsonify({'status': 'error', 'text': f'{e}'})


def parseclients(clients):
    parser = NamesParser()
    total_clients = len(clients)
    current_count = 0
    for item in clients:
        parsename = parser.parse(f'{item["surname"]} {item["name"]} {item["patronymic"]}')
        if parsename["parsed"] == True:
            item["name"] = parsename.get("fn", "")
            item["surname"] = parsename.get("sn", "")
            item["patronymic"] = parsename.get("mn", "")
        if parsename["parsed"] == False:
            item["error"] = True

        current_count += 1
        socketio.emit('splitFIO', {'counter': {'currentcount': current_count, 'totalclients': total_clients}})

    return clients

# простой сплит по символу-разделителю
# def splitByChar(clients):
#     for item in clients:
#         newname = item["name"].split('--')[0]
#         newpatronymic = item['name'].split('--')[-1]
#         print(newpatronymic)
#         item["name"]=newname
#         item["patronymic"]=newpatronymic
#     return clients


# имя оставляет в имени, все остальное - в фамилию
# def simpleSplit(clients):
#     parser = NamesParser()
#     for item in clients:
#         parsename = parser.parse(item["name"])
#         if parsename["parsed"]:
#             item["name"] = parsename.get("fn", "")
#             item["surname"] = parsename.get("sn", "") + " " + parsename.get("mn", "") + item["surname"]
#         else:
#             name_parts = item["name"].split()
#             if len(name_parts) > 0:
#                 item["name"] = name_parts[0]
#                 item["surname"] = " ".join(name_parts[1:]) + item["surname"]
#             item["error"] = True
#     return clients

#
# @fiosplitter.route('/getError', methods=['GET'])
# def getError():
#     try:
#         global newdata
#         global errors
#         errors = []
#         for item in newdata:
#             if 'error' in item:
#                 errors.append(item)
#         return jsonify({'status': 'success', 'text': errors})
#     except Exception as e:
#         return jsonify({'status': 'error', 'text': f'{e}'})
#

@fiosplitter.route('/saveClients', methods=['POST'])
def saveClients():
    try:
        response = request.json
        salon = response["salon_id"]
        login = response["login"]
        password = response["password"]
        if password == "":
            token = login
            runningtime = saveResult(salon, newdata, headers, token)
        else:
            usertoken = getToken(login, password)
            runningtime = saveResult(salon, newdata, headers, usertoken)

        return jsonify({'status': 'success', 'text': f"Успешно сохранено, время выполнения - {runningtime}"})
    except Exception as e:
        return jsonify({'status': 'error', 'text': f'{e}'})


def saveResult(salon, data, headers, usertoken):
    headers['Authorization'] = f'{bearer}, {usertoken}'
    now = datetime.datetime.now()

    total = len(data)  # Общее количество элементов в очереди
    current = 0  # Текущее количество обработанных элементов


    def process_item(item):
        nonlocal current
        url = f'https://api.yclients.com/api/v1/client/{salon}/{item["id"]}'
        payload = json.dumps({
            "id": item['id'],
            "name": item['name'],
            "patronymic": item['patronymic'],
            "phone": item['phone'],
            "surname": item['surname'],
            "birth_date": item['birthday']

        })

        response = requests.request("PUT", url, headers=headers, data=payload)
        # Увеличиваем текущее количество обработанных элементов
        current += 1
        socketio.emit('saveResult', {'counter': {'currentcount': current, 'totalclients': total}})

    with ThreadPoolExecutor() as executor:
        executor.map(process_item, data)
        
        
        
    end = datetime.datetime.now()
    print(f"query is running {end-now}")
    runningtime=end-now
    return runningtime

@fiosplitter.route('/getReport')
def get_report():

    try:
        global newdata
        global errors
        errors = []
        for item in newdata:
            if 'error' in item:
                errors.append(item)

        # Создаем новую workbook и worksheet в Excel
        workbook = Workbook()
        worksheet = workbook.active

        # Записываем заголовки
        headers = ['id', 'Телефон', 'Имя', 'Отчество', 'Фамилия']
        for col, header in enumerate(headers, start=1):
            worksheet.cell(row=1, column=col, value=header)

        # Записываем данные
        for row, item in enumerate(errors, start=2):
            worksheet.cell(row=row, column=1, value=item['id'])
            worksheet.cell(row=row, column=2, value=item['phone'])
            worksheet.cell(row=row, column=3, value=item['name'])
            worksheet.cell(row=row, column=4, value=item['patronymic'])
            worksheet.cell(row=row, column=5, value=item['surname'])

        # Создаем временный буфер для файла
        output = io.BytesIO()

        # Сохраняем workbook во временный буфер
        workbook.save(output)

        # Закрываем workbook
        workbook.close()

        # Подготавливаем response с файлом
        response = make_response(output.getvalue())
        response.headers['Content-Type'] = 'application/vnd.ms-excel'
        response.headers['Content-Disposition'] = 'attachment; filename=report.xlsx'

        return response



    except Exception as e:
        return jsonify({'status': 'error', 'text': f'{e}'})


