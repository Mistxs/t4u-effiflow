import datetime
import json
import os
from concurrent.futures import ThreadPoolExecutor
import logging

import requests
from flask import request, jsonify, Blueprint, current_app, Response
import xlrd
from config import bearer, db_params, headers
from openpyxl import load_workbook
logging.basicConfig(filename='loadclients.log', level=logging.INFO, format='%(asctime)s - %(levelname)s - %(message)s')




loadclients = Blueprint('loadclients', __name__)

clients = []

# Функция для проверки допустимого расширения файла
def allowed_file(filename):
    return '.' in filename and filename.rsplit('.', 1)[1].lower() in {'xls', 'xlsx'}



# Маршрут для загрузки XLS-файла
@loadclients.route('/clients/upload', methods=['POST'])
def upload_file():
    # Проверяем, что файл был отправлен
    print(request.files)
    if 'file' not in request.files:
        return jsonify({'error': 'No file part'}), 400

    file = request.files['file']

    # Проверяем, что файл имеет допустимое расширение
    if file and allowed_file(file.filename):

        upload_folder = current_app.config['UPLOAD_FOLDER']
        # Создаем папку, если она не существует
        os.makedirs(upload_folder, exist_ok=True)

        # Сохраняем файл на сервер
        file_path = os.path.join(upload_folder, file.filename)
        file.save(file_path)

        # Читаем XLS-файл и разбираем его
        data = read_excel_file(file_path)

        # Возвращаем данные в формате JSON
        return jsonify({'data': data})
    else:
        return jsonify({'error': 'Invalid file type'}), 400

@loadclients.route('/clients/saveClients', methods=['POST'])
def saveClients():
    try:
        response = request.json
        salon = response["salon_id"]
        login = response["login"]
        password = response["password"]
        if password == "":
            token = login
            print(clients)
            saveResult(salon, clients, headers, token)
            return jsonify({'status': 'success', 'text': clients})
        else:
            usertoken = getToken(login, password)
            saveResult(salon, clients, headers, usertoken)
            return jsonify({'status': 'success', 'text': clients})
    except Exception as e:
        return jsonify({'status': 'error', 'text': f'{e}'})


def read_excel_file(file_path):
    _, ext = os.path.splitext(file_path)
    ext = ext.lower()

    if ext == '.xls':
        return read_xls_file(file_path)
    elif ext == '.xlsx':
        return read_xlsx_file(file_path)
    else:
        raise ValueError('Unsupported file format')

def read_xls_file(file_path):
    data = []
    workbook = xlrd.open_workbook(file_path)
    sheet = workbook.sheet_by_index(0)  # Предполагаем, что данные находятся на первом листе

    for row in sheet.iter_rows(min_row=2, values_only=True):
        record = {}
        for col_idx, value in enumerate(row, start=1):
            column_name = chr(64 + col_idx)  # Преобразование номера столбца в букву
            record[column_name] = value
        data.append(record)
    return data

def read_xlsx_file(file_path):
    global clients
    workbook = load_workbook(file_path, read_only=True)
    sheet = workbook.active
    for row in sheet.iter_rows(min_row=2, values_only=True):
        name = row[0]
        phone = row[1]
        clients.append({
            'name': name,
            'phone': phone
        })
    return clients

def getToken(login, password):
    url = f"https://api.yclients.com/api/v1/auth"
    payload = json.dumps({
        "login": f"{login}",
        "password": f"{password}"
    })
    response = requests.request("POST", url, headers=headers, data=payload).json()
    if response["success"] == True:
        return response["data"]["user_token"]
    else:
        return f"Ошибка. {response['meta']}"

def saveResult(salon, data, headers, usertoken):
    headers['Authorization'] = f'{bearer}, {usertoken}'
    now = datetime.datetime.now()
    def process_item(item):
        url = f'https://api.yclients.com/api/v1/clients/{salon}'
        payload = json.dumps({
            "name": item.get("name", ""),
            "phone": item.get("phone", ""),
            "patronymic": item.get("patronymic", ""),
            "surname": item.get("surname", "")
        })
        response = requests.request("POST", url, headers=headers, data=payload)
        logging.info(response.text)

    with ThreadPoolExecutor() as executor:
        executor.map(process_item, data)
    end = datetime.datetime.now()
    logging.info(f"query is running {end-now}")
    return data


