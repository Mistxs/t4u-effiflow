import os
from flask import request, jsonify, Blueprint,current_app
import xlrd
from openpyxl import load_workbook



loadvisits = Blueprint('loadvisits', __name__)


# Функция для проверки допустимого расширения файла
def allowed_file(filename):
    return '.' in filename and filename.rsplit('.', 1)[1].lower() in {'xls', 'xlsx'}



# Маршрут для загрузки XLS-файла
@loadvisits.route('/visits/upload', methods=['POST'])
def upload_file():
    # Проверяем, что файл был отправлен
    print(request.files)
    if 'file' not in request.files:
        return jsonify({'error': 'No file part'}), 400

    file = request.files['file']

    # Проверяем, что файл имеет допустимое расширение
    if file and allowed_file(file.filename):

        upload_folder = current_app.config['UPLOAD_FOLDER']
        # Создаем папку, если она не существует
        os.makedirs(upload_folder, exist_ok=True)

        # Сохраняем файл на сервер
        file_path = os.path.join(upload_folder, file.filename)
        file.save(file_path)

        # Читаем XLS-файл и разбираем его
        data = read_excel_file(file_path)

        # Возвращаем данные в формате JSON
        return jsonify({'data': data})
    else:
        return jsonify({'error': 'Invalid file type'}), 400



# Функция для чтения данных из XLS-файла

def read_excel_file(file_path):
    _, ext = os.path.splitext(file_path)
    ext = ext.lower()

    if ext == '.xls':
        return read_xls_file(file_path)
    elif ext == '.xlsx':
        return read_xlsx_file(file_path)
    else:
        raise ValueError('Unsupported file format')

def read_xls_file(file_path):
    data = []
    workbook = xlrd.open_workbook(file_path)
    sheet = workbook.sheet_by_index(0)  # Предполагаем, что данные находятся на первом листе

    for row in sheet.iter_rows(min_row=2, values_only=True):
        record = {}
        for col_idx, value in enumerate(row, start=1):
            column_name = chr(64 + col_idx)  # Преобразование номера столбца в букву
            record[column_name] = value
        data.append(record)

    return data

def read_xlsx_file(file_path):
    data = []
    workbook = load_workbook(file_path, read_only=True)
    sheet = workbook.active
    for row in sheet.iter_rows(min_row=2, values_only=True):
        record = {}
        for col_idx, value in enumerate(row, start=1):
            column_name = chr(64 + col_idx)  # Преобразование номера столбца в букву
            record[column_name] = value
        data.append(record)
    return data