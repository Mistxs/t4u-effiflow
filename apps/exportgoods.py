import asyncio
import base64
import io
import os

from openpyxl.workbook import Workbook

from flask import Blueprint, jsonify, request, send_from_directory, make_response
from config import db_params

import requests

from flask_socketio import SocketIO, emit

import pymysql


connection = pymysql.connect(**db_params)

exportgoods = Blueprint('exportgoods', __name__)
socketio = SocketIO()

@exportgoods.route('/download/<filename>')
def download(filename):
    return send_from_directory(exportgoods.config['UPLOAD_FOLDER'], filename, as_attachment=True)

@exportgoods.route('/exportgoods', methods=["POST"])
def start_Export():
    try:
        chain_id = request.json['chain_id']
        goodsdata = query_to_db(chain_id)
        excel_file_path = createExcel(goodsdata)

        response = make_response(excel_file_path.getvalue())
        response.headers['Content-Type'] = 'application/vnd.ms-excel'
        response.headers['Content-Disposition'] = 'attachment; filename=report.xlsx'

        return response


        # async def generate_and_send_file():
        #     try:
        #         print("Start generate_and_send_file")
        #         goodsdata = query_to_db(chain_id)
        #         excel_file_path = createExcel(goodsdata)
        #
        #         socketio.emit('text', "send text data")
        #
        #         # Перемещаем файл в папку uploads
        #         output_file_path = os.path.join(exportgoods.config['UPLOAD_FOLDER'], 'output.xlsx')
        #         os.rename(excel_file_path, output_file_path)
        #
        #         # Отправляем ссылку на файл на клиент через Socket.IO
        #         socketio.emit('file-ready', {'status': 'success', 'link': '/download/output.xlsx'})
        #
        #         print("File link sent successfully")
        #
        #     except Exception as e:
        #         print(f"Error in generate_and_send_file: {e}")

        # socketio.start_background_task(generate_and_send_file)

        # return jsonify({'status': 'success', "text": "Data processing started"})
    except Exception as e:
        return jsonify({'status': 'error in start_Export', 'text': f'{e}'})

def createExcel(goodsdata):
    try:
        # Создаем новую workbook и worksheet в Excel
        # print(f"start create excel. data: {goodsdata}")
        workbook = Workbook()
        worksheet = workbook.active

        # Записываем заголовки
        headers = ['ID товара', 'ID категории', 'Название категории', 'Родительская категория', 'Название товара',
                   'В каких филиалах присутствует']
        for col, header in enumerate(headers, start=1):
            worksheet.cell(row=1, column=col, value=header)

        # Записываем данные
        for row, item in enumerate(goodsdata, start=2):
            worksheet.cell(row=row, column=1, value=item['id'])
            worksheet.cell(row=row, column=2, value=item['goodcat'])
            worksheet.cell(row=row, column=3, value=item['titlecat'])
            worksheet.cell(row=row, column=4, value=item['pidtitle'])
            worksheet.cell(row=row, column=5, value=item['title'])
            worksheet.cell(row=row, column=6, value=item['salons'])

        # Создаем временный буфер для файла
        output = io.BytesIO()

        # Сохраняем workbook во временный буфер
        workbook.save(output)

        # Закрываем workbook
        workbook.close()

        return output
    except Exception as e:
        return jsonify({'status': 'error', 'text': f'{e}'})


def query_to_db(chain_id):
    try:
        url = f"https://b5898dc6e4bc-8806955829454616363.ngrok-free.app/db_connector/exportgoods?chain={chain_id}"
        # url = f"http://127.0.0.1:5100/db_connector/exportgoods?chain_id={chain_id}"
        response = requests.request("GET", url).json()
        if response["status"] == "success":
            gooddata = response["data"]
            return gooddata
        else:
            raise Exception("Error in db_connector")
    except Exception as e:
        return jsonify({'status': 'error in query_to_db', 'text': f'{e}'})

